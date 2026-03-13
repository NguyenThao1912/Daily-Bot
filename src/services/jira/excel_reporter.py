import os
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import JiraTicket

COLOR = {
    "header_bg": "1E3A5F",
    "header_font": "FFFFFF",
    "epic_bg": "2D6A9F",
    "epic_font": "FFFFFF",
    "subheader_bg": "D6E4F0",
    "bug": "D64545",
    "request": "F39C12",
    "new_feat": "2E9E5B",
    "in_progress": "2F80ED",
    "unknown": "757575",
    "done_bg": "DFF3E3",
    "alert_bg": "FFF3E0",
    "row_alt": "F5F9FF",
    "border": "BDBDBD",
    "todo_status": "FDECEC",
    "progress_status": "EAF3FF",
    "review_status": "FFF4D6",
    "done_status": "E4F7E7",
    "blocked_status": "F8D7DA",
    "reply_yes": "E4F7E7",
    "reply_no": "FDECEC",
    "priority_high": "FDE2E4",
    "priority_medium": "FFF3CD",
    "priority_low": "E2F0D9",
    "link": "1155CC",
}

LABEL_COLORS = {
    "bug": COLOR["bug"],
    "request_feature": COLOR["request"],
    "new_feature": COLOR["new_feat"],
    "in_progress": COLOR["in_progress"],
    "Chưa phân loại": COLOR["unknown"],
}

STATUS_COLORS = {
    "to do": COLOR["todo_status"],
    "open": COLOR["todo_status"],
    "selected for development": COLOR["todo_status"],
    "in progress": COLOR["progress_status"],
    "in development": COLOR["progress_status"],
    "in review": COLOR["review_status"],
    "qa": COLOR["review_status"],
    "testing": COLOR["review_status"],
    "done": COLOR["done_status"],
    "closed": COLOR["done_status"],
    "resolved": COLOR["done_status"],
    "blocked": COLOR["blocked_status"],
}

PRIORITY_COLORS = {
    "highest": COLOR["priority_high"],
    "high": COLOR["priority_high"],
    "medium": COLOR["priority_medium"],
    "low": COLOR["priority_low"],
    "lowest": COLOR["priority_low"],
}


def _thin_border() -> Border:
    side = Side(style="thin", color=COLOR["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _cell(
    ws,
    row: int,
    col: int,
    value: Any = "",
    bold: bool = False,
    bg: Optional[str] = None,
    font_color: str = "000000",
    align: str = "left",
    wrap: bool = False,
    size: int = 10,
):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, size=size, name="Arial")
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _thin_border()
    return cell


def _status_color(status: str) -> Optional[str]:
    normalized = status.strip().lower()
    for key, color in STATUS_COLORS.items():
        if key in normalized:
            return color
    return None


def _priority_color(priority: str) -> Optional[str]:
    return PRIORITY_COLORS.get(priority.strip().lower())


def _ticket_url(ticket_key: str) -> Optional[str]:
    base_url = os.getenv("JIRA_BASE_URL", "").rstrip("/")
    if not base_url:
        return None
    return f"{base_url}/browse/{ticket_key}"


def _group_tickets_by_epic(tickets: List[JiraTicket]) -> List[tuple[str, List[JiraTicket]]]:
    grouped: Dict[str, List[JiraTicket]] = {}
    order: List[str] = []

    for ticket in sorted(
        tickets,
        key=lambda item: (
            item.epic_name or item.epic_key or "ZZZ",
            item.status,
            item.key,
        ),
    ):
        epic_title = ticket.epic_name or ticket.epic_key or "Không thuộc Epic"
        if epic_title not in grouped:
            grouped[epic_title] = []
            order.append(epic_title)
        grouped[epic_title].append(ticket)

    return [(epic_title, grouped[epic_title]) for epic_title in order]


def _sheet_dashboard(wb: Workbook, report: Dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "📊 Dashboard"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = f"🎯 JIRA SCRUM REPORT — {report['project']}"
    title.font = Font(bold=True, size=16, color="FFFFFF", name="Arial")
    title.fill = PatternFill("solid", start_color=COLOR["header_bg"])
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:H2")
    subtitle = ws["A2"]
    subtitle.value = (
        f"Tạo lúc: {report['generated_at']}  |  Tổng tickets: {report['total_tickets']}  |  Epics: {report['epics_count']}"
    )
    subtitle.font = Font(size=10, color="555555", name="Arial")
    subtitle.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    kpis = [
        ("🐛 Bug", report["label_summary"].get("bug", 0), COLOR["bug"]),
        ("📋 Request", report["label_summary"].get("request_feature", 0), COLOR["request"]),
        ("✨ New Feature", report["label_summary"].get("new_feature", 0), COLOR["new_feat"]),
        ("🔄 In Progress", report["label_summary"].get("in_progress", 0), COLOR["in_progress"]),
    ]
    for (label, value, color), col in zip(kpis, [1, 3, 5, 7]):
        ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
        ws.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
        ws.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col + 1)
        ws.cell(row=4, column=col).value = label
        ws.cell(row=4, column=col).font = Font(bold=True, size=11, color=color, name="Arial")
        ws.cell(row=4, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=5, column=col).value = value
        ws.cell(row=5, column=col).font = Font(bold=True, size=28, color=color, name="Arial")
        ws.cell(row=5, column=col).alignment = Alignment(horizontal="center")
        ws.row_dimensions[5].height = 40

    alerts = report.get("scrum_alerts", {})
    ws.merge_cells("A8:H8")
    alert = ws["A8"]
    alert.value = (
        "⚠️  CẢNH BÁO SCRUM MASTER:  "
        f"Stale: {alerts.get('stale_count', 0)}  |  "
        f"Blocked: {alerts.get('blocked_count', 0)}  |  "
        f"Chưa assign: {alerts.get('no_assignee_count', 0)}"
    )
    alert.font = Font(bold=True, size=11, color="BF360C", name="Arial")
    alert.fill = PatternFill("solid", start_color=COLOR["alert_bg"])
    alert.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[8].height = 28

    headers = ["Epic Key", "Epic Name", "Tổng", "Bug", "Request", "New Feature", "In Progress", "Done%"]
    for col, header in enumerate(headers, 1):
        _cell(ws, 10, col, header, bold=True, bg=COLOR["header_bg"], font_color="FFFFFF", align="center")

    row = 11
    for epic_key, epic_data in report["epics"].items():
        bg = COLOR["row_alt"] if row % 2 == 0 else None
        by_label = epic_data["by_label"]
        _cell(ws, row, 1, epic_key, bg=bg)
        _cell(ws, row, 2, epic_data["name"], bg=bg, wrap=True)
        _cell(ws, row, 3, epic_data["total"], bg=bg, align="center")
        _cell(ws, row, 4, by_label.get("bug", 0), bg=COLOR["bug"] if by_label.get("bug") else bg, align="center")
        _cell(ws, row, 5, by_label.get("request_feature", 0), bg=bg, align="center")
        _cell(ws, row, 6, by_label.get("new_feature", 0), bg=bg, align="center")
        _cell(ws, row, 7, by_label.get("in_progress", 0), bg=bg, align="center")
        _cell(
            ws,
            row,
            8,
            f"{epic_data['completion_rate']}%",
            bg=COLOR["done_bg"] if epic_data["completion_rate"] >= 80 else bg,
            align="center",
        )
        row += 1

    for index, width in enumerate([12, 35, 8, 8, 10, 13, 13, 8], 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _sheet_tickets(wb: Workbook, tickets: List[JiraTicket]) -> None:
    ws = wb.create_sheet("📋 Tickets")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "Key",
        "Summary",
        "Epic",
        "Type",
        "Status",
        "Assignee",
        "Dev Reply",
        "Last Dev Reply By",
        "AI Label",
        "Lý do AI",
        "Priority",
        "SP",
        "Updated",
    ]
    for col, header in enumerate(headers, 1):
        _cell(ws, 1, col, header, bold=True, bg=COLOR["header_bg"], font_color="FFFFFF", align="center")
    ws.row_dimensions[1].height = 24

    row = 2
    for epic_title, epic_tickets in _group_tickets_by_epic(tickets):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        epic_header = _cell(
            ws,
            row,
            1,
            f"EPIC: {epic_title} ({len(epic_tickets)} tickets)",
            bold=True,
            bg=COLOR["epic_bg"],
            font_color="FFFFFF",
            size=11,
        )
        epic_header.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22
        row += 1

        for ticket in epic_tickets:
            base_bg = COLOR["row_alt"] if row % 2 == 0 else None
            status_bg = _status_color(ticket.status) or base_bg
            priority_bg = _priority_color(ticket.priority) or base_bg
            label_color = LABEL_COLORS.get(ticket.ai_label, COLOR["unknown"])

            key_cell = _cell(ws, row, 1, ticket.key, bg=base_bg, wrap=True)
            ticket_url = _ticket_url(ticket.key)
            if ticket_url:
                key_cell.hyperlink = ticket_url
                key_cell.font = Font(
                    bold=True,
                    color=COLOR["link"],
                    underline="single",
                    size=10,
                    name="Arial",
                )

            _cell(ws, row, 2, ticket.summary, bg=base_bg, wrap=True)
            _cell(ws, row, 3, ticket.epic_name or ticket.epic_key or "Không thuộc Epic", bg=base_bg, wrap=True)
            _cell(ws, row, 4, ticket.issue_type, bg=base_bg, align="center")
            _cell(ws, row, 5, ticket.status, bg=status_bg, align="center", bold=True)
            _cell(ws, row, 6, ticket.assignee or "—", bg=base_bg, wrap=True)
            _cell(
                ws,
                row,
                7,
                "Đã reply" if ticket.has_dev_reply else "Chưa reply",
                bg=COLOR["reply_yes"] if ticket.has_dev_reply else COLOR["reply_no"],
                align="center",
                bold=True,
            )
            _cell(ws, row, 8, ticket.last_dev_reply_by or "—", bg=base_bg, wrap=True)
            _cell(ws, row, 9, ticket.ai_label or "—", bg=label_color, font_color="FFFFFF", align="center", bold=True)
            _cell(ws, row, 10, ticket.ai_label_reason or "", bg=base_bg, wrap=True)
            _cell(ws, row, 11, ticket.priority or "", bg=priority_bg, align="center", bold=bool(ticket.priority))
            _cell(ws, row, 12, ticket.story_points or "", bg=base_bg, align="center")
            _cell(ws, row, 13, (ticket.updated or ticket.created or "")[:10], bg=base_bg, align="center")
            ws.row_dimensions[row].height = 38
            row += 1

        row += 1

    for index, width in enumerate([16, 42, 26, 14, 18, 20, 14, 18, 16, 42, 12, 8, 12], 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _sheet_alerts(wb: Workbook, tickets: List[JiraTicket], report: Dict[str, Any]) -> None:
    ws = wb.create_sheet("🚨 Scrum Alerts")
    ws.sheet_view.showGridLines = False

    alerts = report.get("scrum_alerts", {})
    headers = ["Key", "Summary", "Status", "Assignee", "Dev Reply", "AI Label"]

    _cell(ws, 1, 1, "⚠️ STALE TICKETS (chưa cập nhật trong nhiều ngày)", bold=True, bg="FF5722", font_color="FFFFFF", size=12)
    ws.merge_cells("A1:F1")

    for col, header in enumerate(headers, 1):
        _cell(ws, 2, col, header, bold=True, bg=COLOR["epic_bg"], font_color="FFFFFF")

    stale_keys = set(alerts.get("stale_keys", []))
    stale_tickets = [ticket for ticket in tickets if ticket.key in stale_keys]
    for row, ticket in enumerate(stale_tickets, 3):
        _cell(ws, row, 1, ticket.key)
        _cell(ws, row, 2, ticket.summary, wrap=True)
        _cell(ws, row, 3, ticket.status, bg=_status_color(ticket.status), align="center", bold=True)
        _cell(ws, row, 4, ticket.assignee or "—")
        _cell(
            ws,
            row,
            5,
            "Đã reply" if ticket.has_dev_reply else "Chưa reply",
            bg=COLOR["reply_yes"] if ticket.has_dev_reply else COLOR["reply_no"],
            align="center",
            bold=True,
        )
        _cell(ws, row, 6, ticket.ai_label or "—", bg=LABEL_COLORS.get(ticket.ai_label, COLOR["unknown"]), font_color="FFFFFF", align="center")

    offset = max(len(stale_tickets) + 5, 6)
    _cell(ws, offset, 1, "🔴 BLOCKED TICKETS", bold=True, bg="B71C1C", font_color="FFFFFF", size=12)
    ws.merge_cells(f"A{offset}:F{offset}")

    for col, header in enumerate(headers, 1):
        _cell(ws, offset + 1, col, header, bold=True, bg=COLOR["epic_bg"], font_color="FFFFFF")

    blocked_keys = set(alerts.get("blocked_keys", []))
    blocked = [ticket for ticket in tickets if ticket.key in blocked_keys]
    for row, ticket in enumerate(blocked, offset + 2):
        _cell(ws, row, 1, ticket.key)
        _cell(ws, row, 2, ticket.summary, wrap=True)
        _cell(ws, row, 3, ticket.status, bg=_status_color(ticket.status), align="center", bold=True)
        _cell(ws, row, 4, ticket.assignee or "—")
        _cell(
            ws,
            row,
            5,
            "Đã reply" if ticket.has_dev_reply else "Chưa reply",
            bg=COLOR["reply_yes"] if ticket.has_dev_reply else COLOR["reply_no"],
            align="center",
            bold=True,
        )
        _cell(ws, row, 6, ticket.ai_label or "—", bg=LABEL_COLORS.get(ticket.ai_label, COLOR["unknown"]), font_color="FFFFFF", align="center")

    for index, width in enumerate([14, 45, 16, 20, 22, 16], 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def export_excel_report(
    tickets: List[JiraTicket],
    report: Dict[str, Any],
    output_path: Optional[str] = None,
) -> str:
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = f"output/jira_report_{report['project']}_{timestamp}.xlsx"

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    workbook = Workbook()
    _sheet_dashboard(workbook, report)
    _sheet_tickets(workbook, tickets)
    _sheet_alerts(workbook, tickets, report)
    workbook.save(output_path)
    return output_path
